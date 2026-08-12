from pathlib import Path

from openpyxl import load_workbook

from biztrip_agent.cli import main
from biztrip_agent.review import generate_review_html
from biztrip_agent.validation import validate_reimbursement


def test_demo_generates_excel_and_review(tmp_path):
    exit_code = main(["demo", "--review", "--output-dir", str(tmp_path)])

    assert exit_code == 0
    workbook_path = next(tmp_path.glob("差旅汇总_demo_*.xlsx"))
    review_path = next(tmp_path.glob("review_*.html"))
    assert workbook_path.exists()
    assert review_path.exists()

    workbook = load_workbook(workbook_path, data_only=True)
    assert workbook.sheetnames == ["报销总览", "费用明细", "按供应商"]
    assert workbook["报销总览"]["D4"].value == "¥ 2,152.50"

    html = review_path.read_text(encoding="utf-8")
    assert "BizTrip Agent 审阅报告" in html
    assert "未发现需要处理的问题" in html
    assert "可以提交" in html


def test_wizard_generates_demo_with_review(tmp_path, monkeypatch):
    answers = iter(["1", str(tmp_path), "y"])
    monkeypatch.setattr("builtins.input", lambda _prompt: next(answers))

    exit_code = main(["wizard"])

    assert exit_code == 0
    assert next(tmp_path.glob("差旅汇总_demo_*.xlsx")).exists()
    assert next(tmp_path.glob("review_*.html")).exists()


def test_review_flags_missing_fields(tmp_path):
    records = [
        {
            "分类": "机票",
            "平台": "去哪儿网",
            "金额": "",
            "日期": "",
            "方法": "规则",
            "附件": "",
        }
    ]

    review_path = generate_review_html(records, [], tmp_path, "测试范围", excel_path=Path("report.xlsx"))
    html = review_path.read_text(encoding="utf-8")

    assert "暂不建议提交" in html
    assert "待补金额" in html
    assert "待补日期" in html
    assert "待补原件" in html
    assert "待确认行程归属" in html


def test_validation_allows_multiple_items_in_one_order():
    records = [
        {
            "分类": "机票",
            "平台": "去哪儿网",
            "金额": 1280.0,
            "日期": "2026-07-10",
            "订单号": "QD001",
            "附件": "flight-a.pdf",
        },
        {
            "分类": "机票",
            "平台": "去哪儿网",
            "金额": 1380.0,
            "日期": "2026-07-10",
            "订单号": "QD001",
            "附件": "flight-b.pdf",
        },
    ]
    trips = [{"records": records}]

    validation = validate_reimbursement(records, trips)

    assert validation["status"] == "ready"
    assert validation["affected_count"] == 0


def test_validation_flags_exact_duplicate_order_items():
    records = [
        {"分类": "机票", "平台": "去哪儿网", "金额": 1280.0, "日期": "2026-07-10", "订单号": "QD001", "附件": "a.pdf"},
        {"分类": "机票", "平台": "去哪儿网", "金额": 1280.0, "日期": "2026-07-10", "订单号": "QD001", "附件": "b.pdf"},
    ]

    validation = validate_reimbursement(records, [{"records": records}])

    assert validation["status"] == "needs_review"
    assert all(any(issue["code"] == "possible_duplicate" for issue in row["issues"]) for row in validation["records"])


def test_validation_flags_conflicting_invoice_numbers():
    records = [
        {"分类": "发票", "供应商": "商店", "金额": 20.0, "日期": "2026-07-10", "发票号码": "INV001", "附件": "a.pdf"},
        {"分类": "发票", "供应商": "商店", "金额": 30.0, "日期": "2026-07-10", "发票号码": "INV001", "附件": "b.pdf"},
    ]

    validation = validate_reimbursement(records, [{"records": records}])

    assert validation["status"] == "needs_review"
    assert all(any(issue["code"] == "identifier_conflict" for issue in row["issues"]) for row in validation["records"])


def test_validation_marks_complete_records_ready():
    record = {
        "分类": "酒店",
        "酒店名称": "华住酒店",
        "金额": 598.0,
        "日期": "2026-07-10",
        "附件": "hotel.pdf",
    }

    validation = validate_reimbursement([record], [{"records": [record]}])

    assert validation["can_submit"] is True
    assert validation["complete_count"] == 1
    assert validation["issue_count"] == 0


def test_validation_blocks_invalid_financial_and_document_values():
    records = [
        {
            "分类": "发票",
            "供应商": "测试商店",
            "金额": float("inf"),
            "日期": "2026-02-30",
            "附件": "../invoice.exe",
        }
    ]

    validation = validate_reimbursement(records, [])
    codes = {issue["code"] for issue in validation["records"][0]["issues"]}

    assert validation["can_submit"] is False
    assert {"invalid_amount", "invalid_date", "invalid_attachment"}.issubset(codes)


def test_validation_blocks_trip_total_dates_and_multiple_assignment():
    record = {
        "记录ID": "R0001",
        "分类": "机票",
        "平台": "去哪儿网",
        "金额": 100.0,
        "日期": "2026-08-05",
        "附件": "flight.pdf",
    }
    trips = [
        {
            "trip_id": 1,
            "start_date": "2026-08-03",
            "end_date": "2026-08-01",
            "records": [record],
            "total": 99.0,
        },
        {
            "trip_id": 2,
            "start_date": "2026-08-01",
            "end_date": "2026-08-02",
            "records": [record],
            "total": 100.0,
        },
    ]

    validation = validate_reimbursement([record], trips)
    codes = {issue["code"] for issue in validation["records"][0]["issues"]}

    assert "multiple_trips" in codes
    assert "invalid_trip_dates" in codes
    assert "trip_total_mismatch" in codes
    assert "date_outside_trip" in codes


def test_validation_checks_physical_attachment_when_directory_is_provided(tmp_path):
    record = {
        "分类": "发票",
        "供应商": "测试商店",
        "金额": 20.0,
        "日期": "2026-08-01",
        "附件": "invoice.pdf",
    }

    missing = validate_reimbursement([record], [], attachment_dir=tmp_path)
    assert "unreadable_attachment" in {
        issue["code"] for issue in missing["records"][0]["issues"]
    }

    (tmp_path / "invoice.pdf").write_bytes(b"%PDF-1.7 content")
    readable = validate_reimbursement([record], [], attachment_dir=tmp_path)
    assert "unreadable_attachment" not in {
        issue["code"] for issue in readable["records"][0]["issues"]
    }


def test_validation_accepts_complete_email_source_evidence(tmp_path):
    record = {
        "分类": "发票",
        "供应商": "测试商店",
        "金额": 20.0,
        "日期": "2026-08-01",
        "附件": "invoice.eml",
    }
    long_received_header = b"Received: " + (b"mail-hop " * 20) + b"\r\n"
    (tmp_path / "invoice.eml").write_bytes(
        long_received_header
        + b"From: invoice@example.com\r\n"
        + b"Subject: Invoice notice\r\n"
        + b"\r\nInvoice details"
    )

    validation = validate_reimbursement([record], [], attachment_dir=tmp_path)

    assert validation["status"] == "ready"
    assert validation["issue_count"] == 0


def test_review_uses_same_physical_attachment_check(tmp_path):
    record = {
        "分类": "发票",
        "供应商": "测试商店",
        "金额": 20.0,
        "日期": "2026-08-01",
        "附件": "missing.pdf",
    }

    review_path = generate_review_html(
        [record],
        [],
        tmp_path,
        "测试范围",
        attachment_dir=tmp_path / "附件",
    )

    page = review_path.read_text(encoding="utf-8")
    assert "暂不建议提交" in page
    assert "原件不存在、为空或无法读取" in page
