import json
from zipfile import ZipFile

from openpyxl import load_workbook

from biztrip_agent.cli import main
from biztrip_agent.results import load_results_json, write_results_json
from phase2.agent_report import (
    _attachments_for_pdf,
    _build_search_query,
    _filter_records_by_requested_dates,
    enrich_records_from_attachments,
    infer_vendor,
)


def test_write_results_json_persists_summary_and_records(tmp_path):
    records = [
        {
            "分类": "机票",
            "金额": 1280.0,
            "日期": "2026-07-10",
            "平台": "去哪儿网",
            "_邮件正文": "不得写入结果的临时证据",
        },
        {"分类": "酒店", "金额": 598.0, "日期": "2026-07-10", "平台": "华住酒店"},
    ]
    trips = [
        {
            "trip_id": 1,
            "destination": "深圳",
            "start_date": "2026-07-10",
            "end_date": "2026-07-10",
            "summary": "深圳出差",
            "records": records,
            "total": 1878.0,
            "method": "规则",
        }
    ]

    path = write_results_json(
        records,
        trips,
        tmp_path,
        "2026-07-01~2026-07-29",
        xlsx_path=tmp_path / "report.xlsx",
        review_path=tmp_path / "review.html",
        agent_task={"schema_version": "biztrip.agent-task.v1", "status": "needs_user_input"},
    )

    assert path.name.startswith("records_")
    assert path.suffix == ".json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    assert payload["schema_version"] == "biztrip.records.v1"
    assert payload["scan_label"] == "2026-07-01~2026-07-29"
    assert payload["summary"]["record_count"] == 2
    assert payload["summary"]["trip_count"] == 1
    assert payload["summary"]["total_amount"] == 1878.0
    assert payload["summary"]["submission_status"] == "needs_review"
    assert payload["summary"]["can_submit"] is False
    assert payload["summary"]["affected_count"] == 2
    assert payload["validation"]["issue_count"] > 0
    assert payload["records"][0]["平台"] == "去哪儿网"
    assert "_邮件正文" not in payload["records"][0]
    assert "不得写入结果的临时证据" not in path.read_text(encoding="utf-8")
    assert payload["agent_task"]["status"] == "needs_user_input"
    assert payload["trips"][0]["destination"] == "深圳"

    loaded = load_results_json(path)
    assert loaded["summary"]["record_count"] == 2


def test_rebuild_generates_excel_and_review_from_json(tmp_path):
    records = [
        {"分类": "机票", "金额": 1280.0, "日期": "2026-07-10", "平台": "去哪儿网", "方法": "规则", "附件": "flight.pdf"},
        {"分类": "酒店", "金额": 598.0, "日期": "2026-07-10", "平台": "华住酒店", "方法": "规则", "附件": "hotel.pdf"},
    ]
    trips = [
        {
            "trip_id": 1,
            "destination": "深圳",
            "start_date": "2026-07-10",
            "end_date": "2026-07-10",
            "summary": "深圳出差",
            "records": records,
            "total": 1878.0,
            "method": "规则",
        }
    ]
    source_dir = tmp_path / "source"
    rebuild_dir = tmp_path / "rebuilt"
    source_json = write_results_json(records, trips, source_dir, "测试范围")

    exit_code = main(["rebuild", str(source_json), "--review", "--output-dir", str(rebuild_dir)])

    assert exit_code == 0
    workbook_path = next(rebuild_dir.glob("差旅汇总_*.xlsx"))
    review_path = next(rebuild_dir.glob("review_*.html"))
    results_path = next(rebuild_dir.glob("records_*.json"))
    assert workbook_path.exists()
    assert review_path.exists()
    assert results_path.exists()

    workbook = load_workbook(workbook_path, data_only=True)
    assert workbook["报销总览"]["D4"].value == "¥ 1,878.00"
    assert "华住酒店" in review_path.read_text(encoding="utf-8")


def test_rebuild_infers_vendor_from_existing_json_context(tmp_path):
    records = [
        {
            "分类": "网约车",
            "金额": 31.09,
            "日期": "2026-07-20",
            "方法": "规则",
            "发件人": "itinerary@ridesharing.amap.com",
            "主题": "高德代驾电子发票",
            "附件": "5_【闪现代驾-31.09元-1个行程】高德代驾电子发票.pdf",
        },
        {
            "分类": "机票",
            "金额": 1688.0,
            "日期": "2026-05-30",
            "方法": "规则",
            "发件人": "去哪儿网 <ticketservice@qunar.com>",
            "主题": "【去哪儿网】机票订单电子报销凭证",
            "附件": "19_2026-05-30 甘孜-成都-机票电子发票-1688.00.pdf",
        },
    ]
    source_json = write_results_json(records, [], tmp_path / "source", "测试范围")
    rebuild_dir = tmp_path / "rebuilt"

    exit_code = main(["rebuild", str(source_json), "--output-dir", str(rebuild_dir)])

    workbook = load_workbook(next(rebuild_dir.glob("差旅汇总_*.xlsx")), data_only=True)
    vendors = {row[0]: row[1] for row in workbook["按供应商"].iter_rows(min_row=3, values_only=True) if row[0]}
    assert exit_code == 0
    assert vendors["去哪儿网"] == 1
    assert vendors["高德"] == 1


def test_infer_vendor_from_invoice_subject_and_sender():
    assert infer_vendor({"主题": "您收到【青羊区乱炒江湖餐厅（个体工商户）】开具的发票"}) == "青羊区乱炒江湖餐厅（个体工商户）"
    assert infer_vendor({"发件人": "盒马财务共享服务中心 <HemaFinSSC@service.freshhema.com>", "主题": "盒马电子发票"}) == "盒马"


def test_enrich_records_backfills_12306_amount_from_archived_zip(tmp_path, monkeypatch):
    output_dir = tmp_path
    attach_dir = output_dir / "附件"
    attach_dir.mkdir()
    zip_path = attach_dir / "6_12306.zip"
    with ZipFile(zip_path, "w") as zf:
        zf.writestr("invoice.pdf", b"pdf-bytes")
    monkeypatch.setattr("phase2.agent_report._pdf_text_from_bytes", lambda _data: "12306 yhN÷:ÿå149.00N")
    records = [
        {
            "分类": "发票",
            "金额": "",
            "日期": "2026年07月15日",
            "供应商": "12306",
            "附件": "6_12306.zip",
        }
    ]

    enrich_records_from_attachments(records, output_dir=output_dir)

    assert records[0]["金额"] == 149.0


def test_search_query_includes_end_date():
    search_cmd, scan_label, limit_recent = _build_search_query("2026-07-01", "2026-07-30", 60)

    assert search_cmd == "SINCE 01-Jul-2026 BEFORE 31-Jul-2026"
    assert scan_label == "2026-07-01~2026-07-30"
    assert limit_recent is False


def test_search_query_limits_recent_only_without_dates():
    search_cmd, scan_label, limit_recent = _build_search_query("", "", 60)

    assert search_cmd == "ALL"
    assert scan_label == "最近60封"
    assert limit_recent is True


def test_final_records_cannot_cross_requested_date_range():
    records = [
        {"日期": "2026-05-31", "主题": "范围前"},
        {"日期": "2026-06-01", "主题": "范围内"},
        {"日期": "2026年07月30日", "主题": "范围内结束"},
        {"日期": "2026-07-31", "主题": "范围后"},
        {"日期": "", "主题": "日期待识别"},
    ]

    filtered = _filter_records_by_requested_dates(records, "2026-06-01", "2026-07-30")

    assert [record["主题"] for record in filtered] == ["范围内", "范围内结束", "日期待识别"]


def test_split_flight_record_keeps_only_its_own_pdf_attachment():
    attachments = ["3_上海-深圳-1280.00.pdf", "3_深圳-北京-980.00.pdf", "3_notice.png"]

    selected = _attachments_for_pdf("深圳-北京-980.00.pdf", attachments)

    assert selected == ["3_深圳-北京-980.00.pdf"]


def test_wizard_rebuilds_from_json(tmp_path, monkeypatch):
    records = [
        {"分类": "机票", "金额": 1280.0, "日期": "2026-07-10", "平台": "去哪儿网", "方法": "规则", "附件": "flight.pdf"},
    ]
    trips = [
        {
            "trip_id": 1,
            "destination": "深圳",
            "start_date": "2026-07-10",
            "end_date": "2026-07-10",
            "summary": "深圳出差",
            "records": records,
            "total": 1280.0,
            "method": "规则",
        }
    ]
    source_json = write_results_json(records, trips, tmp_path / "source", "测试范围")
    rebuild_dir = tmp_path / "rebuilt"
    answers = iter(["2", str(source_json), str(rebuild_dir), "y"])
    monkeypatch.setattr("builtins.input", lambda _prompt: next(answers))

    exit_code = main(["wizard"])

    assert exit_code == 0
    assert next(rebuild_dir.glob("差旅汇总_*.xlsx")).exists()
    assert next(rebuild_dir.glob("review_*.html")).exists()
