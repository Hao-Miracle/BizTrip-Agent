"""BizTrip Agent command-line entry point."""

import argparse
import importlib.util
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

from common.utils import format_chinese_date


PROJECT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = Path(os.getenv("BIZTRIP_DATA_DIR") or PROJECT_DIR)
OUTPUT_DIR = Path(os.getenv("BIZTRIP_OUTPUT_DIR") or DATA_DIR / "output")


def main(argv=None):
    parser = argparse.ArgumentParser(
        prog="biztrip",
        description="Scan travel emails and generate reimbursement reports.",
    )
    parser.add_argument("--version", action="version", version="biztrip-agent 0.1.2")

    subparsers = parser.add_subparsers(dest="command")

    demo_parser = subparsers.add_parser("demo", help="Generate a sample report without reading email.")
    demo_parser.add_argument(
        "--output-dir",
        default=str(OUTPUT_DIR),
        help="Directory for the generated demo report. Defaults to ./output.",
    )
    demo_parser.add_argument("--review", action="store_true", help="Also generate a local HTML review page.")

    subparsers.add_parser("check", help="Check Python version, dependencies, and local config.")
    subparsers.add_parser("init", help="Create .env from .env.example if it does not exist.")
    subparsers.add_parser("wizard", help="Start a guided local workflow.")
    web_parser = subparsers.add_parser("web", help="Start a local web interface.")
    web_parser.add_argument("--host", default="127.0.0.1", help="Host to bind. Defaults to 127.0.0.1.")
    web_parser.add_argument("--port", type=int, default=8765, help="Port to bind. Defaults to 8765.")
    web_parser.add_argument("--no-open", action="store_true", help="Do not open a browser automatically.")
    scan_parser = subparsers.add_parser("scan", help="Run the Agent scanner.")
    scan_parser.add_argument("--start", help="Start date in YYYY-MM-DD format.")
    scan_parser.add_argument("--end", help="End date in YYYY-MM-DD format.")
    scan_parser.add_argument("--count", type=int, default=60, help="Number of recent emails to scan when no date range is set.")
    scan_parser.add_argument("--no-llm", action="store_true", help="Force rule mode even when LLM config is present.")
    scan_parser.add_argument("--review", action="store_true", help="Also generate a local HTML review page.")
    scan_parser.add_argument(
        "--output-dir",
        default=str(OUTPUT_DIR),
        help="Directory for reports and archived attachments. Defaults to ./output.",
    )
    rebuild_parser = subparsers.add_parser("rebuild", help="Regenerate reports from a records JSON file.")
    rebuild_parser.add_argument("json_path", help="Path to records_YYYYMMDD_HHMMSS.json.")
    rebuild_parser.add_argument("--review", action="store_true", help="Also regenerate a local HTML review page.")
    rebuild_parser.add_argument(
        "--output-dir",
        help="Directory for regenerated files. Defaults to the JSON file directory.",
    )
    agent_parser = subparsers.add_parser("agent", help="Machine-readable interface for Agent skills.")
    agent_commands = agent_parser.add_subparsers(dest="agent_command", required=True)
    agent_start = agent_commands.add_parser("start", help="Start a reimbursement task and return JSON.")
    agent_start.add_argument("--start", help="Start date in YYYY-MM-DD format.")
    agent_start.add_argument("--end", help="End date in YYYY-MM-DD format.")
    agent_start.add_argument("--count", type=int, default=60)
    agent_start.add_argument("--no-llm", action="store_true")
    agent_start.add_argument("--output-dir", default=str(OUTPUT_DIR))
    agent_status = agent_commands.add_parser("status", help="Return the latest task as JSON.")
    agent_status.add_argument("--task", help="Specific records JSON path.")
    agent_status.add_argument("--output-dir", default=str(OUTPUT_DIR))
    agent_answer = agent_commands.add_parser("answer", help="Submit user confirmations from JSON.")
    agent_answer.add_argument("--task", required=True, help="Current records JSON path.")
    agent_answer.add_argument("--answers-file", required=True, help="UTF-8 JSON file containing answers.")
    agent_answer.add_argument("--output-dir")

    args = parser.parse_args(argv)
    command = args.command or "demo"

    if command == "demo":
        return demo(Path(args.output_dir), review=args.review)
    if command == "check":
        return check()
    if command == "init":
        return init_config()
    if command == "wizard":
        return wizard()
    if command == "web":
        return web(args)
    if command == "scan":
        return scan(args)
    if command == "rebuild":
        return rebuild(args)
    if command == "agent":
        return agent_command(args)

    parser.print_help()
    return 2


def agent_command(args):
    """Run the stable JSON-only interface used by thin Agent skills."""
    from biztrip_agent.agent_interface import answer_task, start_task, task_status

    if args.agent_command == "start":
        if args.count < 1 or any(
            value and not _is_date(value) for value in (args.start, args.end)
        ):
            payload = {
                "schema_version": "biztrip.agent-interface.v1",
                "operation": "start",
                "ok": False,
                "status": "failed",
                "error": {"code": "invalid_range", "message": "日期必须使用 YYYY-MM-DD，count 必须大于 0。"},
                "next_action": "inspect_error",
            }
        else:
            payload = start_task(args.start, args.end, args.count, args.no_llm, args.output_dir)
    elif args.agent_command == "status":
        payload = task_status(args.task, args.output_dir)
    else:
        try:
            answers = json.loads(Path(args.answers_file).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            payload = {
                "schema_version": "biztrip.agent-interface.v1",
                "operation": "answer",
                "ok": False,
                "status": "failed",
                "error": {"code": "answers_unreadable", "message": str(exc)[:500]},
                "next_action": "inspect_error",
            }
        else:
            payload = answer_task(args.task, answers, args.output_dir)
    print(json.dumps(payload, ensure_ascii=False))
    return 0 if payload.get("ok") else 1


def demo(output_dir, review=False):
    """Generate a local Excel report from fictional records."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Missing dependency: openpyxl")
        print("Install with: pip install -e .")
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)
    records = _demo_records()
    trips = _demo_trips(records)
    total = sum(record["金额"] for record in records)

    wb = Workbook()
    blue = "1A56DB"
    dark = "1F2937"
    header_fill = PatternFill(start_color=blue, end_color=blue, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
    title_font = Font(bold=True, color=dark, size=14, name="微软雅黑")
    body_font = Font(size=10, name="微软雅黑")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws = wb.active
    ws.title = "报销总览"
    ws.merge_cells("A1:F1")
    ws["A1"] = "差旅费用报销汇总单 · Demo"
    ws["A1"].font = title_font
    ws.merge_cells("A2:F2")
    ws["A2"] = f"生成日期：{format_chinese_date()} | 示例数据 | 出差相关：{len(records)} 条"
    ws["A2"].font = Font(size=9, color="6B7280", name="微软雅黑")
    ws.merge_cells("A4:C4")
    ws["A4"] = "可报销总额"
    ws["A4"].font = Font(size=12, color="6B7280", name="微软雅黑")
    ws.merge_cells("D4:F4")
    ws["D4"] = f"¥ {total:,.2f}"
    ws["D4"].font = Font(bold=True, size=22, color=blue, name="微软雅黑")
    ws["D4"].alignment = Alignment(horizontal="right", vertical="center")

    row = 6
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "出差行程汇总"
    ws[f"A{row}"].font = Font(bold=True, size=11, color=dark, name="微软雅黑")
    row += 1
    _write_header(ws, row, ["行程编号", "目的地", "起止日期", "订单数", "金额合计", "摘要"], header_font, header_fill, center, border)
    for trip in trips:
        row += 1
        values = [
            f"Trip #{trip['trip_id']}",
            trip["destination"],
            f"{trip['start_date']} ~ {trip['end_date']}",
            len(trip["records"]),
            trip["total"],
            trip["summary"],
        ]
        _write_row(ws, row, values, body_font, center, border)

    row += 2
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "按类别汇总"
    ws[f"A{row}"].font = Font(bold=True, size=11, color=dark, name="微软雅黑")
    row += 1
    _write_header(ws, row, ["类别", "笔数", "金额(元)", "占比", "", ""], header_font, header_fill, center, border)
    for category, amount in _category_totals(records):
        row += 1
        count = sum(1 for record in records if record["分类"] == category)
        pct = f"{amount / total * 100:.1f}%" if total else "0.0%"
        _write_row(ws, row, [category, count, amount, pct, "", ""], body_font, center, border)
    row += 1
    _write_row(ws, row, ["合计", len(records), total, "100%", "", ""], Font(bold=True, color=blue, size=11, name="微软雅黑"), center, border)

    for col, width in enumerate([12, 12, 16, 12, 14, 34], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws2 = wb.create_sheet("费用明细")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "差旅费用明细表"
    ws2["A1"].font = title_font
    headers = ["序号", "日期", "类别", "供应商/平台", "金额(元)", "出发地→目的地", "方法", "附件"]
    _write_header(ws2, 2, headers, header_font, header_fill, center, border)
    for index, record in enumerate(sorted(records, key=lambda item: item["日期"]), 1):
        route = f"{record.get('出发地', '')}→{record.get('目的地', '')}" if record.get("出发地") else "-"
        values = [
            index,
            record["日期"],
            record["分类"],
            record.get("平台") or record.get("供应商") or "-",
            record["金额"],
            route,
            record["方法"],
            record.get("附件") or "-",
        ]
        _write_row(ws2, index + 2, values, body_font, left, border)
    for col, width in enumerate([6, 14, 10, 18, 12, 20, 12, 28], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws3 = wb.create_sheet("按供应商")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "按供应商/平台统计"
    ws3["A1"].font = title_font
    _write_header(ws3, 2, ["供应商", "笔数", "金额(元)", "占比"], header_font, header_fill, center, border)
    for index, (vendor, amount, count) in enumerate(_vendor_totals(records), 1):
        pct = f"{amount / total * 100:.1f}%" if total else "0.0%"
        _write_row(ws3, index + 2, [vendor, count, amount, pct], body_font, center, border)
    for col, width in enumerate([20, 10, 14, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = width

    from biztrip_agent.results import unique_output_path

    report_path = unique_output_path(output_dir, "差旅汇总_demo", ".xlsx")
    wb.save(report_path)

    print("Demo report generated.")
    print(f"Records: {len(records)}")
    print(f"Total: ¥{total:,.2f}")
    print(f"Excel: {report_path}")
    if review:
        from biztrip_agent.review import generate_review_html

        review_path = generate_review_html(
            records,
            trips,
            output_dir,
            scan_label="示例数据",
            excel_path=report_path,
        )
        print(f"Review: {review_path}")
    return 0


def check():
    """Check local readiness without contacting email servers."""
    ok = True
    print(f"Python: {sys.version.split()[0]}")
    if sys.version_info < (3, 8):
        print("FAIL: Python 3.8+ is required.")
        ok = False

    for module_name, package_name, required in [
        ("dotenv", "python-dotenv", True),
        ("PyPDF2", "PyPDF2", True),
        ("openpyxl", "openpyxl", True),
        ("openai", "openai", False),
    ]:
        available = importlib.util.find_spec(module_name) is not None
        status = "OK" if available else ("MISSING" if required else "OPTIONAL")
        print(f"{package_name}: {status}")
        if required and not available:
            ok = False

    env_path = Path(os.getenv("BIZTRIP_ENV_PATH") or DATA_DIR / ".env")
    if env_path.exists():
        print(".env: found")
    else:
        print(".env: not found; run `biztrip init` before scanning email.")

    return 0 if ok else 1


def init_config():
    """Create a starter .env file without reading any secrets."""
    env_path = Path(os.getenv("BIZTRIP_ENV_PATH") or DATA_DIR / ".env")
    example_path = env_path.parent / ".env.example"
    if env_path.exists():
        print(f".env already exists: {env_path}")
        return 0
    if not example_path.exists():
        print(f"Missing template: {example_path}")
        return 1
    shutil.copyfile(example_path, env_path)
    print(f"Created: {env_path}")
    print("Edit EMAIL_ACCOUNT and EMAIL_PASSWORD before running `biztrip scan`.")
    return 0


def wizard():
    """Guide non-technical users through the safest local workflows."""
    print("BizTrip Agent wizard")
    print("1. Generate a demo report")
    print("2. Rebuild reports from an existing records JSON")
    print("3. Check local environment")
    print("4. Exit")

    choice = input("Choose 1-4 [1]: ").strip() or "1"
    if choice == "1":
        output_dir = input(f"Output directory [{OUTPUT_DIR}]: ").strip() or str(OUTPUT_DIR)
        review = _confirm("Also create a review page?", default=True)
        return demo(Path(output_dir), review=review)
    if choice == "2":
        json_path = input("Path to records_YYYYMMDD_HHMMSS.json: ").strip()
        if not json_path:
            print("A JSON path is required.")
            return 2
        output_dir = input("Output directory [same as JSON file]: ").strip()
        review = _confirm("Also create a review page?", default=True)
        args = argparse.Namespace(json_path=json_path, output_dir=output_dir or None, review=review)
        return rebuild(args)
    if choice == "3":
        return check()
    if choice == "4":
        print("No changes made.")
        return 0

    print("Please choose 1, 2, 3, or 4.")
    return 2


def web(args):
    """Run the local web UI."""
    if args.port < 1 or args.port > 65535:
        print("--port must be between 1 and 65535.")
        return 2
    try:
        from biztrip_agent.web import run_server
    except ImportError as exc:
        print(f"Unable to load web UI: {exc}")
        return 1
    return run_server(host=args.host, port=args.port, open_browser=not args.no_open)


def scan(args):
    """Run the scanner, interactively by default or non-interactively with options."""
    if args.count < 1:
        print("--count must be greater than 0.")
        return 2
    for label, value in [("--start", args.start), ("--end", args.end)]:
        if value and not _is_date(value):
            print(f"{label} must use YYYY-MM-DD format.")
            return 2

    try:
        from phase2.agent_report import main as agent_main
    except ImportError as exc:
        print(f"Unable to load scanner: {exc}")
        return 1
    interactive = not any([args.start, args.end, args.count != 60, args.no_llm, args.output_dir != str(OUTPUT_DIR)])
    result = agent_main(
        start=args.start,
        end=args.end,
        count=args.count,
        no_llm=args.no_llm,
        output_dir=args.output_dir,
        interactive=interactive,
        review=args.review,
    )
    if result and result.get("review_path"):
        print(f"Review: {result['review_path']}")
    return 0


def rebuild(args):
    """Regenerate Excel and optional review HTML from a saved JSON result."""
    try:
        from biztrip_agent.results import load_results_json, write_results_json
        from phase2.agent_report import _generate_excel, enrich_records_from_attachments
        from phase2.llm_aggregate import aggregate_trips
    except ImportError as exc:
        print(f"Unable to load rebuild tools: {exc}")
        return 1

    json_path = Path(args.json_path)
    if not json_path.exists():
        print(f"JSON file not found: {json_path}")
        return 2

    try:
        payload = load_results_json(json_path)
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        print(f"Unable to read results JSON: {exc}")
        return 2

    output_dir = Path(args.output_dir) if args.output_dir else json_path.parent
    records = payload["records"]
    enrich_records_from_attachments(records, output_dir=json_path.parent)
    trips = aggregate_trips(records, use_llm=False)
    total_amount = sum(record.get("金额", 0) or 0 for record in records)
    scan_label = payload.get("scan_label") or "JSON rebuild"
    xlsx_path = _generate_excel(
        records,
        trips,
        total_amount,
        scan_label,
        output_dir=str(output_dir),
        use_llm=False,
    )

    review_path = None
    if args.review:
        from biztrip_agent.review import generate_review_html

        review_path = generate_review_html(records, trips, output_dir, scan_label, excel_path=xlsx_path)

    results_path = write_results_json(records, trips, output_dir, scan_label, xlsx_path=xlsx_path, review_path=review_path)
    print("Reports rebuilt from JSON.")
    print(f"Excel: {xlsx_path}")
    if review_path:
        print(f"Review: {review_path}")
    print(f"JSON: {results_path}")
    return 0


def _is_date(value):
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return False
    return True


def _confirm(prompt, default=False):
    suffix = "Y/n" if default else "y/N"
    answer = input(f"{prompt} [{suffix}]: ").strip().lower()
    if not answer:
        return default
    return answer in {"y", "yes"}


def _write_header(ws, row, headers, font, fill, alignment, border):
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border


def _write_row(ws, row, values, font, alignment, border):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _demo_records():
    return [
        {
            "分类": "机票",
            "平台": "去哪儿网",
            "金额": 1280.0,
            "日期": "2026-07-10",
            "出发地": "上海",
            "目的地": "深圳",
            "方法": "Demo",
            "附件": "demo_flight_invoice.pdf",
        },
        {
            "分类": "酒店",
            "平台": "华住酒店",
            "金额": 598.0,
            "日期": "2026-07-10",
            "目的地": "深圳",
            "方法": "Demo",
            "附件": "demo_hotel_invoice.pdf",
        },
        {
            "分类": "网约车",
            "平台": "滴滴出行",
            "金额": 86.5,
            "日期": "2026-07-10",
            "出发地": "深圳宝安机场",
            "目的地": "南山区酒店",
            "方法": "Demo",
            "附件": "demo_taxi_1.pdf",
        },
        {
            "分类": "网约车",
            "平台": "高德打车",
            "金额": 52.0,
            "日期": "2026-07-11",
            "出发地": "南山区酒店",
            "目的地": "客户办公室",
            "方法": "Demo",
            "附件": "demo_taxi_2.pdf",
        },
        {
            "分类": "发票",
            "平台": "智慧发票",
            "金额": 136.0,
            "日期": "2026-07-11",
            "供应商": "深圳示例餐饮有限公司",
            "方法": "Demo",
            "附件": "demo_meal_invoice.pdf",
        },
    ]


def _demo_trips(records):
    total = sum(record["金额"] for record in records)
    return [
        {
            "trip_id": 1,
            "destination": "深圳",
            "start_date": "2026-07-10",
            "end_date": "2026-07-11",
            "total": total,
            "summary": "上海到深圳客户拜访，含机票、酒店、打车和餐饮发票。",
            "records": records,
        }
    ]


def _category_totals(records):
    totals = {}
    for record in records:
        totals[record["分类"]] = totals.get(record["分类"], 0) + record["金额"]
    return sorted(totals.items(), key=lambda item: -item[1])


def _vendor_totals(records):
    totals = {}
    for record in records:
        vendor = record.get("供应商") or record.get("平台") or "其他"
        if vendor not in totals:
            totals[vendor] = {"count": 0, "amount": 0}
        totals[vendor]["count"] += 1
        totals[vendor]["amount"] += record["金额"]
    return [
        (vendor, data["amount"], data["count"])
        for vendor, data in sorted(totals.items(), key=lambda item: -item[1]["amount"])
    ]


if __name__ == "__main__":
    raise SystemExit(main())
